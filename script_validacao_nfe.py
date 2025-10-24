#!/usr/bin/env python3
"""
Sistema de Validação de NFe's - Versão 3.0 FINAL
Desenvolvido para Windows

LÓGICA COMPLETA IMPLEMENTADA:
1. Agrupamento correto (individual vs agrupado com vírgula)
2. Divergentes Agrupadas vs Divergentes Simples (lógica correta)
3. Aba de Resumo com todos os cálculos e validação
4. Tratamento de Cancelamento e Estornos
5. Múltiplos ZIPs
6. Relatório Excel completo

Autor: Claude
Data: 2025-10-24
Versão: 3.0
"""

import xml.etree.ElementTree as ET
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date
from typing import List, Dict, Optional, Tuple, Set
from dataclasses import dataclass, field
from collections import defaultdict
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import sys

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

class Config:
    """Configurações do sistema"""
    CNPJ_EMPRESA = "17122471000175"
    DATA_FECHAMENTO = date(2025, 9, 30)
    PERIODO_MAXIMO_DIAS = 60
    TOLERANCIA_VALOR = Decimal("1.01")
    
    COLUNAS = {
        'numero_nfe': 'AS',
        'data_abastecimento': 'D',
        'cnpj_posto': 'H',
        'cnpj_empresa': 'J',
        'valor_boleto': 'AO',
        'postergado': 'AR'
    }

# ============================================================================
# MODELOS DE DADOS
# ============================================================================

@dataclass
class NFe:
    """Nota Fiscal Eletrônica extraída do XML"""
    numero: str
    data_emissao: date
    cnpj_emitente: str
    cnpj_destinatario: str
    valor: Decimal
    xml_path: str
    
    def __post_init__(self):
        self.cnpj_emitente = re.sub(r'\D', '', self.cnpj_emitente)
        self.cnpj_destinatario = re.sub(r'\D', '', self.cnpj_destinatario)

@dataclass
class TransacaoOriginal:
    """Transação original da planilha (1 linha)"""
    numeros_nfe: Set[str]  # Pode ter múltiplos: {"103576", "103577"}
    data_abastecimento: date
    cnpj_posto: str
    cnpj_empresa: str
    valor_boleto: Decimal
    postergado: str
    linha: int
    texto_original: str  # Texto original da célula
    eh_cancelamento: bool = False
    eh_estorno: bool = False
    
    def __post_init__(self):
        self.cnpj_posto = re.sub(r'\D', '', self.cnpj_posto)
        self.cnpj_empresa = re.sub(r'\D', '', self.cnpj_empresa)
    
    @property
    def eh_agrupamento(self) -> bool:
        """True se tem múltiplos números (vírgula)"""
        return len(self.numeros_nfe) > 1

@dataclass
class GrupoNFe:
    """Grupo de uma NFe específica para confrontamento"""
    numero_nfe: str
    data_abastecimento: date
    cnpj_posto: str
    cnpj_empresa: str
    valor_planilha: Decimal
    postergado: str
    transacoes_individuais: List[TransacaoOriginal]
    transacoes_agrupadas: List[TransacaoOriginal]
    linhas: List[int]
    
    @property
    def tem_individual(self) -> bool:
        return len(self.transacoes_individuais) > 0
    
    @property
    def tem_agrupado(self) -> bool:
        return len(self.transacoes_agrupadas) > 0
    
    @property
    def tipo_origem(self) -> str:
        """Retorna: INDIVIDUAL, AGRUPADO ou AMBOS"""
        if self.tem_individual and self.tem_agrupado:
            return "AMBOS"
        elif self.tem_individual:
            return "INDIVIDUAL"
        elif self.tem_agrupado:
            return "AGRUPADO"
        return "NENHUM"

@dataclass
class Resultado:
    """Resultado do confrontamento"""
    tipo: str  # IDENTICA, DIVERGENTE, DIVERGENTE_AGRUPADA, NAO_ENCONTRADA, DESCONSIDERADA
    grupo: GrupoNFe
    nfe: Optional[NFe]
    diferenca: Decimal
    dias_postergados: int
    motivo: str
    grupo_divergencia_id: Optional[int] = None  # Para agrupamentos de múltiplas NFes

@dataclass
class Resumo:
    """Resumo final dos cálculos"""
    valor_do_boleto: Decimal
    identicas: Decimal
    agrupadas: Decimal
    divergentes: Decimal
    desconsideradas: Decimal
    nao_encontradas: Decimal
    cancelamento: Decimal
    estornos: Decimal
    total: Decimal
    valor_a_pagar: Decimal
    diferenca_validacao: Decimal

# ============================================================================
# UTILITÁRIOS
# ============================================================================

def limpar_cnpj(cnpj: str) -> str:
    """Remove formatação do CNPJ"""
    if cnpj is None:
        return ""
    return re.sub(r'\D', '', str(cnpj))

def formatar_cnpj(cnpj: str) -> str:
    """Formata CNPJ: 12345678000190 -> 12.345.678/0001-90"""
    cnpj = limpar_cnpj(cnpj)
    if len(cnpj) != 14:
        return cnpj
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"

def converter_valor(valor) -> Decimal:
    """Converte valor para Decimal com precisão"""
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, Decimal):
        return valor
    
    if isinstance(valor, (int, float)):
        return Decimal(str(valor)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    valor_str = str(valor).strip().replace('R$', '').replace('$', '').strip()
    
    if ',' in valor_str and '.' in valor_str:
        ultima_virgula = valor_str.rfind(',')
        ultimo_ponto = valor_str.rfind('.')
        
        if ultima_virgula > ultimo_ponto:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        else:
            valor_str = valor_str.replace(',', '')
    elif ',' in valor_str:
        valor_str = valor_str.replace(',', '.')
    
    try:
        return Decimal(valor_str).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except:
        return Decimal('0')

def converter_data(valor) -> Optional[date]:
    """Converte valor para date"""
    if valor is None:
        return None
    
    if isinstance(valor, date) and not isinstance(valor, datetime):
        return valor
    
    if isinstance(valor, datetime):
        return valor.date()
    
    valor_str = str(valor).strip()
    formatos = ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"]
    
    for formato in formatos:
        try:
            dt = datetime.strptime(valor_str, formato)
            return dt.date()
        except:
            continue
    
    return None

def extrair_numeros_nfe(texto: str) -> Tuple[Set[str], bool, bool]:
    """
    Extrai números de NFe, detecta cancelamento e estorno.
    
    Retorna: (set de números, eh_cancelamento, eh_estorno)
    
    Exemplos:
    "NFe16184" -> ({"16184"}, False, False)
    "NFe103576, NFe103577" -> ({"103576", "103577"}, False, False)
    "Cancelado" -> (set(), True, False)
    "Estorno" -> (set(), False, True)
    """
    if not texto:
        return (set(), False, False)
    
    texto_str = str(texto).strip()
    texto_lower = texto_str.lower()
    
    # Verificar cancelamento/estorno
    eh_cancelamento = 'cancelado' in texto_lower or 'cancelamento' in texto_lower
    eh_estorno = 'estorno' in texto_lower
    
    if eh_cancelamento or eh_estorno:
        return (set(), eh_cancelamento, eh_estorno)
    
    # Extrair números
    texto_limpo = re.sub(r'[^\d,]', '', texto_str)
    numeros = texto_limpo.split(',')
    numeros_set = {n.strip() for n in numeros if n.strip()}
    
    return (numeros_set, False, False)

# ============================================================================
# PARSER DE XML
# ============================================================================

def extrair_dados_xml_from_bytes(xml_data: bytes, xml_name: str) -> Optional[NFe]:
    """Extrai dados de um XML em bytes"""
    try:
        root = ET.fromstring(xml_data)
        
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}')[1]
        
        numero_nfe = root.find(".//ide/nNF")
        data_emissao_elem = root.find(".//ide/dhEmi")
        if data_emissao_elem is None:
            data_emissao_elem = root.find(".//ide/dEmi")
        cnpj_emitente = root.find(".//emit/CNPJ")
        cnpj_destinatario = root.find(".//dest/CNPJ")
        valor_nfe = root.find(".//total/ICMSTot/vNF")
        
        if not all([
            numero_nfe is not None and numero_nfe.text,
            data_emissao_elem is not None and data_emissao_elem.text,
            cnpj_emitente is not None and cnpj_emitente.text,
            cnpj_destinatario is not None and cnpj_destinatario.text,
            valor_nfe is not None and valor_nfe.text
        ]):
            return None
        
        data_str = data_emissao_elem.text.split('T')[0]
        data_emissao = datetime.strptime(data_str, "%Y-%m-%d").date()
        
        return NFe(
            numero=numero_nfe.text,
            data_emissao=data_emissao,
            cnpj_emitente=cnpj_emitente.text,
            cnpj_destinatario=cnpj_destinatario.text,
            valor=Decimal(valor_nfe.text),
            xml_path=xml_name
        )
    except:
        return None

def carregar_xmls_de_multiplos_zips(pasta_zips: Path, mostrar_progresso: bool = True) -> List[NFe]:
    """Carrega XMLs de múltiplos arquivos ZIP"""
    nfes = []
    
    if pasta_zips.is_file() and pasta_zips.suffix == '.zip':
        zip_files = [pasta_zips]
    else:
        zip_files = list(pasta_zips.glob("*.zip"))
    
    if not zip_files:
        print(f"⚠️  Nenhum arquivo ZIP encontrado em {pasta_zips}")
        return []
    
    print(f"\n📂 Encontrados {len(zip_files)} arquivo(s) ZIP:")
    for zf in zip_files:
        print(f"   - {zf.name}")
    
    for zip_idx, zip_path in enumerate(zip_files, 1):
        print(f"\n📦 Processando ZIP {zip_idx}/{len(zip_files)}: {zip_path.name}")
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                xml_files = [f for f in zip_ref.namelist() if f.lower().endswith('.xml')]
                print(f"   Encontrados: {len(xml_files)} arquivos XML")
                
                count_antes = len(nfes)
                
                for i, xml_name in enumerate(xml_files, 1):
                    if mostrar_progresso and i % 1000 == 0:
                        print(f"   Processando: {i}/{len(xml_files)} ({i/len(xml_files)*100:.1f}%)")
                    
                    try:
                        xml_data = zip_ref.read(xml_name)
                        nfe = extrair_dados_xml_from_bytes(xml_data, xml_name)
                        if nfe:
                            nfes.append(nfe)
                    except:
                        continue
                
                print(f"   ✅ Carregadas: {len(nfes) - count_antes} NFe's deste ZIP")
        
        except Exception as e:
            print(f"   ❌ Erro ao processar ZIP: {e}")
            continue
    
    print(f"\n✅ Total de NFe's carregadas: {len(nfes)}")
    return nfes

# ============================================================================
# PARSER DE PLANILHA
# ============================================================================

def carregar_transacoes_da_planilha(planilha_path: Path, config: Config) -> Tuple[List[TransacaoOriginal], Decimal]:
    """
    Carrega transações da planilha e calcula valor total do boleto.
    
    Retorna: (lista de transações, valor_total_boleto)
    """
    print(f"\n📊 Carregando planilha {planilha_path.name}...")
    
    wb = openpyxl.load_workbook(planilha_path, data_only=True)
    ws = wb.active
    
    print(f"   Aba: {ws.title}")
    print(f"   Linhas: {ws.max_row}")
    
    col_indices = {}
    for nome, letra in config.COLUNAS.items():
        col_indices[nome] = openpyxl.utils.column_index_from_string(letra)
    
    transacoes = []
    erros = 0
    valor_total_boleto = Decimal('0')
    
    for linha in range(2, ws.max_row + 1):
        try:
            numero_nfe_texto = ws.cell(linha, col_indices['numero_nfe']).value
            data_abast = ws.cell(linha, col_indices['data_abastecimento']).value
            cnpj_posto = ws.cell(linha, col_indices['cnpj_posto']).value
            cnpj_empresa = ws.cell(linha, col_indices['cnpj_empresa']).value
            valor = ws.cell(linha, col_indices['valor_boleto']).value
            postergado = ws.cell(linha, col_indices['postergado']).value
            
            if not numero_nfe_texto or not data_abast:
                continue
            
            data_convertida = converter_data(data_abast)
            if not data_convertida:
                erros += 1
                continue
            
            valor_convertido = converter_valor(valor)
            valor_total_boleto += valor_convertido
            
            # Extrair números de NFe
            numeros_nfe, eh_cancelamento, eh_estorno = extrair_numeros_nfe(numero_nfe_texto)
            
            transacao = TransacaoOriginal(
                numeros_nfe=numeros_nfe,
                data_abastecimento=data_convertida,
                cnpj_posto=str(cnpj_posto) if cnpj_posto else "",
                cnpj_empresa=str(cnpj_empresa) if cnpj_empresa else config.CNPJ_EMPRESA,
                valor_boleto=valor_convertido,
                postergado=str(postergado) if postergado else "Não",
                linha=linha,
                texto_original=str(numero_nfe_texto),
                eh_cancelamento=eh_cancelamento,
                eh_estorno=eh_estorno
            )
            
            transacoes.append(transacao)
        
        except Exception as e:
            erros += 1
            continue
    
    print(f"   ✅ Carregadas: {len(transacoes)} transações")
    print(f"   💰 Valor total do boleto: R$ {valor_total_boleto:,.2f}")
    if erros > 0:
        print(f"   ⚠️  Erros/Ignoradas: {erros}")
    
    return transacoes, valor_total_boleto

# ============================================================================
# AGRUPAMENTO E ÍNDICE
# ============================================================================

def criar_grupos_nfe(transacoes: List[TransacaoOriginal]) -> Dict[str, GrupoNFe]:
    """
    Cria grupos por número de NFe.
    
    LÓGICA CRÍTICA:
    - Separa transações individuais de agrupadas
    - Calcula valor_planilha conforme regras:
      1. Se tem individual: soma só individuais
      2. Se só agrupado: primeiro número pega valor, outros pegam 0
    """
    print(f"\n🔗 Criando grupos por número de NFe...")
    
    # Índice: numero_nfe -> {individuais: [], agrupadas: []}
    indice = defaultdict(lambda: {
        'individuais': [],
        'agrupadas': []
    })
    
    # Separar transações
    for trans in transacoes:
        # Ignorar cancelamentos e estornos
        if trans.eh_cancelamento or trans.eh_estorno:
            continue
        
        for numero in trans.numeros_nfe:
            if trans.eh_agrupamento:
                indice[numero]['agrupadas'].append(trans)
            else:
                indice[numero]['individuais'].append(trans)
    
    # Criar grupos
    grupos = {}
    
    for numero, dados in indice.items():
        individuais = dados['individuais']
        agrupadas = dados['agrupadas']
        
        # Calcular valor_planilha
        if len(individuais) > 0:
            # Tem individual: soma só individuais
            valor_planilha = sum(t.valor_boleto for t in individuais)
            todas_trans = individuais + agrupadas
            linhas = [t.linha for t in individuais]
        
        elif len(agrupadas) > 0:
            # Só agrupado: verificar se é primeiro número
            primeira_trans_agrupada = agrupadas[0]
            numeros_na_ordem = sorted(primeira_trans_agrupada.numeros_nfe)
            
            if numero == numeros_na_ordem[0]:
                # É o primeiro: pega valor total dos agrupamentos
                valor_planilha = sum(t.valor_boleto for t in agrupadas)
            else:
                # Não é o primeiro: valor = 0
                valor_planilha = Decimal('0')
            
            todas_trans = agrupadas
            linhas = [t.linha for t in agrupadas]
        
        else:
            continue
        
        # Pegar dados da primeira transação
        primeira = individuais[0] if individuais else agrupadas[0]
        
        grupo = GrupoNFe(
            numero_nfe=numero,
            data_abastecimento=primeira.data_abastecimento,
            cnpj_posto=primeira.cnpj_posto,
            cnpj_empresa=primeira.cnpj_empresa,
            valor_planilha=valor_planilha,
            postergado=primeira.postergado,
            transacoes_individuais=individuais,
            transacoes_agrupadas=agrupadas,
            linhas=linhas
        )
        
        grupos[numero] = grupo
    
    # Estatísticas
    grupos_individuais = [g for g in grupos.values() if g.tipo_origem == "INDIVIDUAL"]
    grupos_agrupados = [g for g in grupos.values() if g.tipo_origem == "AGRUPADO"]
    grupos_ambos = [g for g in grupos.values() if g.tipo_origem == "AMBOS"]
    
    print(f"   Total de grupos: {len(grupos)}")
    print(f"   Grupos individuais: {len(grupos_individuais)}")
    print(f"   Grupos só agrupados: {len(grupos_agrupados)}")
    print(f"   Grupos ambos (individual+agrupado): {len(grupos_ambos)}")
    
    if grupos_ambos:
        print(f"\n   📋 Exemplos de grupos com ambos:")
        for i, grupo in enumerate(grupos_ambos[:5], 1):
            print(f"      {i}. NFe {grupo.numero_nfe}: "
                  f"{len(grupo.transacoes_individuais)} individuais + "
                  f"{len(grupo.transacoes_agrupadas)} agrupadas → "
                  f"Valor: R$ {grupo.valor_planilha:.2f}")
    
    return grupos

# ============================================================================
# MOTOR DE CONFRONTAMENTO
# ============================================================================

def calcular_dias_postergados(data_abastecimento: date, data_fechamento: date) -> int:
    """Calcula dias postergados"""
    delta = data_fechamento - data_abastecimento
    return max(0, delta.days)

def confrontar_grupo(grupo: GrupoNFe, nfes_index: Dict[str, NFe], config: Config) -> Resultado:
    """
    Confronta um grupo com as NFe's disponíveis.
    
    LÓGICA DE CLASSIFICAÇÃO:
    - IDENTICA: diferença <= tolerância
    - DIVERGENTE_AGRUPADA: só tem agrupado (tipo_origem == AGRUPADO)
    - DIVERGENTE: tem individual (tipo_origem == INDIVIDUAL ou AMBOS)
    - NAO_ENCONTRADA: XML não existe
    - DESCONSIDERADA: > 60 dias
    """
    
    # 1. Buscar NFe
    nfe = nfes_index.get(grupo.numero_nfe)
    
    if nfe is None:
        dias = calcular_dias_postergados(grupo.data_abastecimento, config.DATA_FECHAMENTO)
        return Resultado(
            tipo="NAO_ENCONTRADA",
            grupo=grupo,
            nfe=None,
            diferenca=grupo.valor_planilha,
            dias_postergados=dias,
            motivo="XML não encontrado"
        )
    
    # 2. Calcular dias postergados
    dias = calcular_dias_postergados(grupo.data_abastecimento, config.DATA_FECHAMENTO)
    
    # 3. Verificar se está desconsiderada
    if dias > config.PERIODO_MAXIMO_DIAS:
        return Resultado(
            tipo="DESCONSIDERADA",
            grupo=grupo,
            nfe=nfe,
            diferenca=abs(grupo.valor_planilha - nfe.valor),
            dias_postergados=dias,
            motivo=f"Postergada: {dias} dias > {config.PERIODO_MAXIMO_DIAS} dias"
        )
    
    # 4. Calcular diferença
    diferenca = abs(grupo.valor_planilha - nfe.valor)
    
    # 5. Classificar
    if diferenca <= config.TOLERANCIA_VALOR:
        tipo = "IDENTICA"
        motivo = "Match exato"
    else:
        # Divergente: escolher tipo
        if grupo.tipo_origem == "AGRUPADO":
            tipo = "DIVERGENTE_AGRUPADA"
            motivo = f"Divergência (só agrupado): R$ {diferenca:.2f}"
        else:
            # INDIVIDUAL ou AMBOS
            tipo = "DIVERGENTE"
            motivo = f"Divergência: R$ {diferenca:.2f}"
    
    return Resultado(
        tipo=tipo,
        grupo=grupo,
        nfe=nfe,
        diferenca=diferenca,
        dias_postergados=dias,
        motivo=motivo
    )

def processar_confrontamento(grupos: Dict[str, GrupoNFe], nfes: List[NFe],
                            config: Config) -> Dict[str, List[Resultado]]:
    """Processa o confrontamento completo"""
    
    print(f"\n🔍 Iniciando confrontamento...")
    print(f"   Grupos: {len(grupos)}")
    print(f"   NFe's: {len(nfes)}")
    print(f"   Data de fechamento: {config.DATA_FECHAMENTO.strftime('%d/%m/%Y')}")
    print(f"   Tolerância: R$ {config.TOLERANCIA_VALOR:.2f}")
    
    # Criar índice de NFe's
    nfes_index = {nfe.numero: nfe for nfe in nfes}
    print(f"   Índice criado: {len(nfes_index)} chaves únicas")
    
    # Processar
    resultados = {
        'IDENTICA': [],
        'DIVERGENTE': [],
        'DIVERGENTE_AGRUPADA': [],
        'NAO_ENCONTRADA': [],
        'DESCONSIDERADA': []
    }
    
    total = len(grupos)
    for i, grupo in enumerate(grupos.values(), 1):
        if i % 1000 == 0:
            print(f"   Processando: {i}/{total} ({i/total*100:.1f}%)")
        
        resultado = confrontar_grupo(grupo, nfes_index, config)
        resultados[resultado.tipo].append(resultado)
    
    print(f"   ✅ Confrontamento concluído!")
    
    return resultados

# ============================================================================
# DETECÇÃO DE GRUPOS DE DIVERGÊNCIAS AGRUPADAS
# ============================================================================

def detectar_grupos_divergentes_agrupadas(divergentes_agrupadas: List[Resultado]) -> Dict[int, List[Resultado]]:
    """
    Detecta grupos: múltiplas NFes que vieram do mesmo agrupamento original.
    
    Exemplo: "NFe103576, NFe103577" gera grupo 1 com ambas.
    """
    print(f"\n🔗 Detectando grupos de Divergentes Agrupadas...")
    
    # Agrupar por transação agrupada original
    grupos = defaultdict(list)
    
    for resultado in divergentes_agrupadas:
        # Pegar primeira transação agrupada (todas do grupo vêm da mesma)
        if resultado.grupo.transacoes_agrupadas:
            primeira_trans = resultado.grupo.transacoes_agrupadas[0]
            # Usar linha como chave do grupo
            chave = primeira_trans.linha
            grupos[chave].append(resultado)
    
    # Numerar grupos
    grupos_numerados = {}
    grupo_id = 1
    
    for chave, resultados_grupo in grupos.items():
        if len(resultados_grupo) >= 2:
            # É um grupo válido
            for r in resultados_grupo:
                r.grupo_divergencia_id = grupo_id
            grupos_numerados[grupo_id] = resultados_grupo
            grupo_id += 1
        else:
            # Só 1 nota, ainda assim atribuir grupo
            resultados_grupo[0].grupo_divergencia_id = grupo_id
            grupos_numerados[grupo_id] = resultados_grupo
            grupo_id += 1
    
    if grupos_numerados:
        print(f"   ✅ Encontrados: {len(grupos_numerados)} grupos")
        for gid, notas in list(grupos_numerados.items())[:5]:
            numeros = [r.grupo.numero_nfe for r in notas]
            print(f"      Grupo {gid}: NFes {', '.join(numeros)}")
    else:
        print(f"   ℹ️  Nenhum grupo detectado")
    
    return grupos_numerados

# ============================================================================
# CÁLCULO DE CANCELAMENTOS E ESTORNOS
# ============================================================================

def calcular_cancelamentos_estornos(transacoes: List[TransacaoOriginal]) -> Tuple[Decimal, Decimal]:
    """
    Calcula totais de cancelamentos e estornos.
    
    Retorna: (total_cancelamento, total_estorno)
    """
    cancelamento = sum(t.valor_boleto for t in transacoes if t.eh_cancelamento)
    estorno = sum(t.valor_boleto for t in transacoes if t.eh_estorno)
    
    # Estornos geralmente são negativos
    if estorno > 0:
        estorno = -estorno
    
    return cancelamento, estorno

# ============================================================================
# GERAÇÃO DE RESUMO
# ============================================================================

def gerar_resumo(resultados: Dict[str, List[Resultado]], 
                valor_total_boleto: Decimal,
                cancelamento: Decimal,
                estorno: Decimal) -> Resumo:
    """Gera resumo final com todos os cálculos"""
    
    # Somar valores por categoria
    identicas = sum(r.grupo.valor_planilha for r in resultados['IDENTICA'])
    
    divergentes_agrupadas = sum(r.grupo.valor_planilha for r in resultados['DIVERGENTE_AGRUPADA'])
    
    divergentes = sum(r.grupo.valor_planilha for r in resultados['DIVERGENTE'])
    
    desconsideradas = sum(r.nfe.valor for r in resultados['DESCONSIDERADA'] if r.nfe)
    
    nao_encontradas = sum(r.grupo.valor_planilha for r in resultados['NAO_ENCONTRADA'])
    
    # Calcular total
    total = identicas + divergentes_agrupadas + divergentes + desconsideradas - nao_encontradas
    
    # Valor a pagar
    valor_a_pagar = total - desconsideradas
    
    # Diferença de validação
    diferenca_validacao = abs(valor_a_pagar - valor_total_boleto)
    
    return Resumo(
        valor_do_boleto=valor_total_boleto,
        identicas=identicas,
        agrupadas=divergentes_agrupadas,
        divergentes=divergentes,
        desconsideradas=desconsideradas,
        nao_encontradas=nao_encontradas,
        cancelamento=cancelamento,
        estornos=estorno,
        total=total,
        valor_a_pagar=valor_a_pagar,
        diferenca_validacao=diferenca_validacao
    )

# ============================================================================
# RELATÓRIO TERMINAL
# ============================================================================

def gerar_relatorio_terminal(resultados: Dict[str, List[Resultado]], 
                            resumo: Resumo,
                            grupos_div_agrup: Dict[int, List[Resultado]]):
    """Gera relatório no terminal"""
    
    total = sum(len(lista) for lista in resultados.values())
    
    print("\n" + "="*100)
    print("📊 RELATÓRIO DE CONFRONTAMENTO")
    print("="*100)
    
    print(f"\n📈 RESUMO:")
    print(f"   Total processado: {total:,}")
    print(f"   ✅ Idênticas: {len(resultados['IDENTICA']):,} ({len(resultados['IDENTICA'])/total*100:.1f}%)")
    print(f"   ⚠️  Divergentes Simples: {len(resultados['DIVERGENTE']):,} ({len(resultados['DIVERGENTE'])/total*100:.1f}%)")
    print(f"   🔗 Divergentes Agrupadas: {len(grupos_div_agrup):,} grupos ({len(resultados['DIVERGENTE_AGRUPADA'])} notas)")
    print(f"   ❌ Não encontradas: {len(resultados['NAO_ENCONTRADA']):,} ({len(resultados['NAO_ENCONTRADA'])/total*100:.1f}%)")
    print(f"   🕐 Desconsideradas: {len(resultados['DESCONSIDERADA']):,} ({len(resultados['DESCONSIDERADA'])/total*100:.1f}%)")
    
    print(f"\n💰 VALORES TOTAIS:")
    print(f"   Valor do boleto:    R$ {resumo.valor_do_boleto:>15,.2f}")
    print(f"   Idênticas:          R$ {resumo.identicas:>15,.2f}")
    print(f"   Agrupadas:          R$ {resumo.agrupadas:>15,.2f}")
    print(f"   Divergentes:        R$ {resumo.divergentes:>15,.2f}")
    print(f"   Desconsideradas:    R$ {resumo.desconsideradas:>15,.2f}")
    print(f"   Não encontradas:    R$ {resumo.nao_encontradas:>15,.2f}")
    print(f"   Cancelamento:       R$ {resumo.cancelamento:>15,.2f}")
    print(f"   Estornos:           R$ {resumo.estornos:>15,.2f}")
    print(f"   " + "-" * 70)
    print(f"   Total:              R$ {resumo.total:>15,.2f}")
    print(f"   Valor a pagar:      R$ {resumo.valor_a_pagar:>15,.2f}")
    print(f"   " + "=" * 70)
    print(f"   Diferença:          R$ {resumo.diferenca_validacao:>15,.2f}")
    
    if resumo.diferenca_validacao < 1000:
        print(f"   ✅ Validação OK!")
    else:
        print(f"   ⚠️  Diferença alta - verificar!")

# ============================================================================
# GERAÇÃO DE RELATÓRIO EXCEL
# ============================================================================

def criar_relatorio_excel(resultados: Dict[str, List[Resultado]], 
                         grupos_div_agrup: Dict[int, List[Resultado]],
                         resumo: Resumo,
                         transacoes: List[TransacaoOriginal],
                         output_path: Path,
                         config: Config):
    """Cria relatório Excel completo"""
    
    print(f"\n📄 Gerando relatório Excel...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================================================
    # ABA 1: RESUMO
    # ========================================================================
    print(f"   Criando aba: Resumo...")
    ws_resumo = wb.create_sheet("Resumo", 0)
    
    ws_resumo['A1'] = "RESUMO DO CONFRONTAMENTO"
    ws_resumo['A1'].font = Font(bold=True, size=14)
    
    linha = 3
    dados_resumo = [
        ("Valor do boleto", resumo.valor_do_boleto),
        ("Idênticas", resumo.identicas),
        ("Agrupadas", resumo.agrupadas),
        ("Divergentes", resumo.divergentes),
        ("Desconsideradas", resumo.desconsideradas),
        ("Não encontradas", resumo.nao_encontradas),
        ("Cancelamento", resumo.cancelamento),
        ("Estornos", resumo.estornos),
        ("", ""),
        ("Total", resumo.total),
        ("Valor a pagar", resumo.valor_a_pagar),
        ("", ""),
        ("Diferença (Validação)", resumo.diferenca_validacao),
    ]
    
    for label, valor in dados_resumo:
        ws_resumo[f'A{linha}'] = label
        if label:
            ws_resumo[f'A{linha}'].font = Font(bold=True)
        if isinstance(valor, Decimal):
            ws_resumo[f'B{linha}'] = float(valor)
            ws_resumo[f'B{linha}'].number_format = 'R$ #,##0.00'
        linha += 1
    
    # Validação
    ws_resumo[f'A{linha}'] = "Status"
    ws_resumo[f'A{linha}'].font = Font(bold=True)
    if resumo.diferenca_validacao < 1000:
        ws_resumo[f'B{linha}'] = "✅ OK"
        ws_resumo[f'B{linha}'].font = Font(color="00B050", bold=True)
    else:
        ws_resumo[f'B{linha}'] = "⚠️ VERIFICAR"
        ws_resumo[f'B{linha}'].font = Font(color="FF0000", bold=True)
    
    ws_resumo.column_dimensions['A'].width = 25
    ws_resumo.column_dimensions['B'].width = 20
    
    # ========================================================================
    # ABA 2: NFe IDÊNTICAS
    # ========================================================================
    print(f"   Criando aba: NFe Idênticas...")
    ws_identicas = wb.create_sheet("NFe Idênticas")
    
    headers = ["N° Nota", "Data Abastecimento", "CNPJ Emitente", "CNPJ Destinatário", 
               "Valor Cobrado", "Diferença", "Linhas Planilha"]
    
    ws_identicas.append(headers)
    for cell in ws_identicas[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for resultado in resultados['IDENTICA']:
        linhas = ', '.join(str(l) for l in resultado.grupo.linhas)
        ws_identicas.append([
            resultado.grupo.numero_nfe,
            resultado.grupo.data_abastecimento.strftime('%d/%m/%Y'),
            formatar_cnpj(resultado.nfe.cnpj_emitente if resultado.nfe else ""),
            formatar_cnpj(resultado.nfe.cnpj_destinatario if resultado.nfe else ""),
            float(resultado.grupo.valor_planilha),
            float(resultado.diferenca),
            linhas
        ])
    
    # ========================================================================
    # ABA 3: NFe DIVERGENTES AGRUPADAS
    # ========================================================================
    print(f"   Criando aba: NFe Divergentes Agrupadas...")
    ws_div_agrup = wb.create_sheet("NFe Divergentes Agrupadas")
    
    headers = ["N° Nota", "Grupo", "Data Abastecimento", "CNPJ Emitente", "CNPJ Destinatário",
               "Valor Receita", "Valor Planilha", "Diferença"]
    
    ws_div_agrup.append(headers)
    for cell in ws_div_agrup[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Ordenar por grupo
    div_agrup_ordenadas = sorted(resultados['DIVERGENTE_AGRUPADA'], 
                                 key=lambda r: r.grupo_divergencia_id or 0)
    
    for resultado in div_agrup_ordenadas:
        ws_div_agrup.append([
            resultado.grupo.numero_nfe,
            resultado.grupo_divergencia_id or "",
            resultado.grupo.data_abastecimento.strftime('%d/%m/%Y'),
            formatar_cnpj(resultado.nfe.cnpj_emitente if resultado.nfe else ""),
            formatar_cnpj(resultado.nfe.cnpj_destinatario if resultado.nfe else ""),
            float(resultado.nfe.valor) if resultado.nfe else 0,
            float(resultado.grupo.valor_planilha),
            float(resultado.diferenca)
        ])
    
    # ========================================================================
    # ABA 4: NFe DIVERGENTES
    # ========================================================================
    print(f"   Criando aba: NFe Divergentes...")
    ws_divergentes = wb.create_sheet("NFe Divergentes")
    
    headers = ["N° Nota", "Data Abastecimento", "CNPJ Emitente", "CNPJ Destinatário",
               "Valor XML", "Valor Planilha", "Diferença", "Linhas Planilha"]
    
    ws_divergentes.append(headers)
    for cell in ws_divergentes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for resultado in resultados['DIVERGENTE']:
        linhas = ', '.join(str(l) for l in resultado.grupo.linhas)
        ws_divergentes.append([
            resultado.grupo.numero_nfe,
            resultado.grupo.data_abastecimento.strftime('%d/%m/%Y'),
            formatar_cnpj(resultado.nfe.cnpj_emitente if resultado.nfe else ""),
            formatar_cnpj(resultado.nfe.cnpj_destinatario if resultado.nfe else ""),
            float(resultado.nfe.valor) if resultado.nfe else 0,
            float(resultado.grupo.valor_planilha),
            float(resultado.diferenca),
            linhas
        ])
    
    # ========================================================================
    # ABA 5: NFe NÃO ENCONTRADAS
    # ========================================================================
    print(f"   Criando aba: NFe Não Encontradas...")
    ws_nao_enc = wb.create_sheet("NFe Não Encontradas")
    
    headers = ["N° Nota", "Data Abastecimento", "CNPJ Posto", "CNPJ Empresa",
               "Valor Planilha", "Linhas Planilha"]
    
    ws_nao_enc.append(headers)
    for cell in ws_nao_enc[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for resultado in resultados['NAO_ENCONTRADA']:
        linhas = ', '.join(str(l) for l in resultado.grupo.linhas)
        ws_nao_enc.append([
            resultado.grupo.numero_nfe,
            resultado.grupo.data_abastecimento.strftime('%d/%m/%Y'),
            formatar_cnpj(resultado.grupo.cnpj_posto),
            formatar_cnpj(resultado.grupo.cnpj_empresa),
            float(resultado.grupo.valor_planilha),
            linhas
        ])
    
    # ========================================================================
    # ABA 6: NFe DESCONSIDERADAS
    # ========================================================================
    print(f"   Criando aba: NFe Desconsideradas...")
    ws_descons = wb.create_sheet("NFe Desconsideradas")
    
    headers = ["N° Nota", "Data Abastecimento", "CNPJ Emitente", "CNPJ Destinatário",
               "Valor Nfe", "Dias Postergados", "Motivo"]
    
    ws_descons.append(headers)
    for cell in ws_descons[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for resultado in resultados['DESCONSIDERADA']:
        ws_descons.append([
            resultado.grupo.numero_nfe,
            resultado.grupo.data_abastecimento.strftime('%d/%m/%Y'),
            formatar_cnpj(resultado.nfe.cnpj_emitente if resultado.nfe else ""),
            formatar_cnpj(resultado.nfe.cnpj_destinatario if resultado.nfe else ""),
            float(resultado.nfe.valor) if resultado.nfe else 0,
            resultado.dias_postergados,
            resultado.motivo
        ])
    
    # ========================================================================
    # ABA 7: CANCELAMENTOS/ESTORNOS
    # ========================================================================
    print(f"   Criando aba: Cancelamentos Estornos...")
    ws_cancel = wb.create_sheet("Cancelamentos Estornos")
    
    headers = ["Tipo", "Texto Original", "Data Abastecimento", "CNPJ Posto", "Valor", "Linha"]
    
    ws_cancel.append(headers)
    for cell in ws_cancel[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for trans in transacoes:
        if trans.eh_cancelamento or trans.eh_estorno:
            tipo = "Cancelamento" if trans.eh_cancelamento else "Estorno"
            ws_cancel.append([
                tipo,
                trans.texto_original,
                trans.data_abastecimento.strftime('%d/%m/%Y'),
                formatar_cnpj(trans.cnpj_posto),
                float(trans.valor_boleto),
                trans.linha
            ])
    
    # ========================================================================
    # AJUSTAR LARGURA DAS COLUNAS
    # ========================================================================
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar
    wb.save(output_path)
    print(f"   ✅ Relatório salvo: {output_path}")

# ============================================================================
# MAIN
# ============================================================================

def main():
    """Função principal"""
    print("="*100)
    print("🔍 SISTEMA DE VALIDAÇÃO DE NFe's - VERSÃO 3.0 FINAL")
    print("="*100)
    
    config = Config()
    
    # CAMINHOS - ADAPTAR PARA SEU AMBIENTE
    pasta_zips = Path(r"C:/Users/kauan.brasileiro/Desktop/validador-nfe-completo/validador-nfe/uploads")
    planilha_path = Path(r"C:/Users/kauan.brasileiro/Desktop/validador-nfe-completo/validador-nfe/uploads/Relatório_NFe_s_-_16_a_30_SET.xlsx")
    output_path = Path(r"C:/Users/kauan.brasileiro/Desktop/validador-nfe-completo/validador-nfe/outputs/Relatório_Final_V3.xlsx")
    
    # 1. Carregar XMLs
    nfes = carregar_xmls_de_multiplos_zips(pasta_zips)
    
    if not nfes:
        print("❌ Nenhum XML válido encontrado!")
        return
    
    # 2. Carregar planilha
    transacoes, valor_total_boleto = carregar_transacoes_da_planilha(planilha_path, config)
    
    if not transacoes:
        print("❌ Nenhuma transação válida encontrada!")
        return
    
    # 3. Criar grupos por NFe
    grupos = criar_grupos_nfe(transacoes)
    
    if not grupos:
        print("❌ Nenhum grupo criado!")
        return
    
    # 4. Confrontamento
    resultados = processar_confrontamento(grupos, nfes, config)
    
    # 5. Detectar grupos de divergentes agrupadas
    grupos_div_agrup = detectar_grupos_divergentes_agrupadas(resultados['DIVERGENTE_AGRUPADA'])
    
    # 6. Calcular cancelamentos e estornos
    cancelamento, estorno = calcular_cancelamentos_estornos(transacoes)
    
    # 7. Gerar resumo
    resumo = gerar_resumo(resultados, valor_total_boleto, cancelamento, estorno)
    
    # 8. Relatórios
    gerar_relatorio_terminal(resultados, resumo, grupos_div_agrup)
    
    # 9. Excel
    output_path.parent.mkdir(parents=True, exist_ok=True)
    criar_relatorio_excel(resultados, grupos_div_agrup, resumo, transacoes, output_path, config)
    
    print("\n" + "="*100)
    print("✅ PROCESSAMENTO CONCLUÍDO!")
    print("="*100)
    print(f"\n📊 Relatório Excel: {output_path}")
    print(f"💾 Tamanho do arquivo: {output_path.stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    main()